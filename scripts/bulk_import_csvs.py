"""Bulk-import all Garmin CSV/XLSX files from a folder into PostgreSQL.

Handles both .csv and .xlsx files. For .xlsx files, converts each sheet to
an in-memory CSV string and runs it through the same importer. Fully
idempotent — safe to run multiple times; the DB dedup logic prevents duplicates.

Usage:
    .venv\\Scripts\\python.exe -m scripts.bulk_import_csvs <folder>

Example:
    .venv\\Scripts\\python.exe -m scripts.bulk_import_csvs data\\CSVs
"""
from __future__ import annotations

import io
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from backend import csv_importer, db

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


def _import_xlsx(path: Path) -> int:
    """Convert each sheet of an xlsx to CSV and run through the importer."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Serialise to an in-memory CSV string
        buf = io.StringIO()
        import csv as csv_mod
        writer = csv_mod.writer(buf)
        for row in rows:
            writer.writerow(["" if v is None else str(v) for v in row])
        buf.seek(0)
        content = buf.read()
        # Feed straight into the CSV parser
        delimiter = csv_importer._detect_delimiter(content[:4096])
        reader = csv_mod.DictReader(io.StringIO(content), delimiter=delimiter)
        headers = reader.fieldnames or []
        csv_rows = list(reader)
        summaries = csv_importer.parse_rows(csv_rows, headers)
        for s in summaries:
            db.upsert_activity(
                s,
                {"_source": "xlsx_import", "_file": path.name, "_sheet": sheet_name},
            )
        total += len(summaries)
        logger.info("  Sheet '%s': %d running activities parsed", sheet_name, len(summaries))
    wb.close()
    return total


def main(argv: list[str] | None = None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    if not argv:
        print("Usage: python -m scripts.bulk_import_csvs <folder-with-csvs>")
        return 2

    folder = Path(argv[0])
    if not folder.is_dir():
        print(f"ERROR: '{folder}' is not a directory.")
        return 1

    files = sorted(folder.iterdir())
    csv_files = [f for f in files if f.suffix.lower() == ".csv"]
    xlsx_files = [f for f in files if f.suffix.lower() in (".xlsx", ".xls")]
    skipped = [f for f in files if f.suffix.lower() not in (".csv", ".xlsx", ".xls")]

    print(f"\nFound {len(csv_files)} CSV file(s) and {len(xlsx_files)} Excel file(s).")
    if skipped:
        print(f"Skipping {len(skipped)} unrecognised file(s): {[f.name for f in skipped]}")
    print()

    before = db.count_activities()
    grand_total = 0
    errors = []

    all_files = csv_files + xlsx_files
    for i, path in enumerate(all_files, 1):
        print(f"[{i}/{len(all_files)}] {path.name} ...", end=" ", flush=True)
        try:
            if path.suffix.lower() == ".csv":
                n = csv_importer.import_csv_file(path)
            else:
                n = _import_xlsx(path)
            grand_total += n
            print(f"{n} runs imported")
        except Exception as e:
            print(f"ERROR — {e}")
            errors.append((path.name, str(e)))

    after = db.count_activities()
    net_new = after - before

    print()
    print("=" * 50)
    print(f"  Done.")
    print(f"  Runs parsed across all files : {grand_total}")
    print(f"  DB before                    : {before}")
    print(f"  DB after                     : {after}")
    print(f"  Net new rows added           : {net_new}")
    if errors:
        print(f"\n  Errors ({len(errors)}):")
        for name, err in errors:
            print(f"    {name}: {err}")
    print("=" * 50)
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
